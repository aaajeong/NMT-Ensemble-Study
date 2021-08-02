#!/usr/bin/env python
# coding: utf-8

# In[1]:


import torch
import openpyxl
import numpy as np
from tqe import TQE


# In[ ]:


wb = openpyxl.load_workbook('Survival_Translate.xlsx')
sheet = wb['sheet']
# source : 번역 원문
# target : 기계 번역 문장

for i in range(2, 1002):
    target = []
    source = []
    candidate = sheet.cell(row = i, column = 4).value.capitalize()
    reference = sheet.cell(row = i, column = 3).value
    print(i, ':', candidate)

    # Translation Quality Estimator (QE)
    # https://github.com/theRay07/Translation-Quality-Estimator
    target.append(candidate)
    source.append(reference)
    model = TQE('LaBSE')
    cos_sim_values = model.fit(source, target)
    sheet.cell(row = i, column = 5).value = cos_sim_values[0]
    
wb.save('Survival_Translate.xlsx')

# # lang_1 = ["what are you doing", "what is your name"]
# # lang_2 = ["तुम क्या कर रहे हो", "तुम्हारा नाम क्या है"]
# lang_1 = ["what are you doing here"]
# lang_2 = ["तुम यहां क्या कर रहे होो"]

# # lang_1 = ["they were not there .", "i never work with him ."]
# # lang_2 = ["Ellos no estaban allá.", "Jamás trabajé con él."]

# model = TQE('LaBSE')
# cos_sim_values = model.fit(lang_1, lang_2)
# print(cos_sim_values)


# In[3]:


wb = openpyxl.load_workbook('Survival_Translate.xlsx')
sheet = wb['sheet']
qe = []


for i in range(2, 1002):
    qe.append(float(sheet.cell(row = i, column = 5).value))

    
qe = np.array(qe)

sheet.cell(row = 1002, column = 5).value = qe.mean()


wb.save('Survival_Translate.xlsx')

